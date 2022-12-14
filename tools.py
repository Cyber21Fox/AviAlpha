# -*- coding: utf8 -*-
import os
import re
import selen
import time
import config
import selen
import math
import traceback
import openpyxl
import pandas as pd
import logging as log
from qt.change_option import Ui_Widget
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from PySide6.QtCore import QObject, QRunnable, Qt, QThreadPool
from PySide6 import QtGui, QtCore
from PySide6.QtCore import QObject, QThread, Signal, Slot
from multiprocessing.dummy import Pool as ThreadPool
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
import magic
from googleapiclient.discovery import build
from google.oauth2 import service_account
from googleapiclient.http import MediaFileUpload

from PySide6.QtWidgets import QApplication, QWidget, QMessageBox, QTableWidgetItem, QTabWidget, QFileDialog, QHeaderView, QMainWindow
SCOPES = ['https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = 'D:\Develop\Project.py\AvitoSelen\keys\projectforavito-2b0dee3fea57.json'
credentials = service_account.Credentials.from_service_account_file(
                        SERVICE_ACCOUNT_FILE, scopes=SCOPES)


class ChangeOption(QWidget):
    def __init__(self):
        super(ChangeOption, self).__init__()
        self.ch = Ui_Widget()
        self.ch.setupUi(self)
        self.ch.pushButton.clicked.connect(self.return_combobox)

    def return_combobox(self):
        a = self.ch.comboBox.currentText()
        self.close()

    def set_combobox(self, item_list):
        self.ch.comboBox.addItems(item_list)


class Tools(QFileDialog):
    def __init__(self, parent):
        super(Tools, self).__init__()#(parent)
        self.ui = parent
        self.data = config.update_ad
        self.browser = selen.Browser(self.ui)
        self.change = ChangeOption()

    def add_image_drive(self):
        try:
            file, _ = QFileDialog.getOpenFileNames(self, 'Open File', './', "Image (*.png *.jpg *jpeg)")  # file ???????? ???? ??????????
            self.ui.tableWidget_image_drive.setColumnCount(10)
            self.ui.tableWidget_image_drive.setRowCount(math.ceil(len(file) / 10))
            self.ui.progressBar.setValue(10)
            self.ui.tableWidget_image_drive.setIconSize(QtCore.QSize(100, 150))
            column = 0
            self.ui.label_image_drive.setText(f"?????????????? ???????????????? : {len(file)}")
            self.ui.progressBar.setValue(30)
            for image_path in file:
                item = QTableWidgetItem()
                self.ui.tableWidget_image_drive.setItem(0, column, item)
                item.setIcon(QtGui.QIcon(image_path))
                column += 1
            self.ui.progressBar.setValue(40)
            log.info(f"???????? ??????????????!")
            self.ui.tableWidget_image_drive.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
            self.ui.tableWidget_image_drive.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
            self.ui.tableWidget_image_drive.horizontalHeader().setMinimumSectionSize(0)
            list_link = []
            drive_service = build('drive', 'v3', credentials=credentials)
            self.ui.progressBar.setValue(60)
            for image in file:
                basename = os.path.basename(image)
                mime = magic.from_file(image, mime=True)
                file_metadata = {"name": basename, "mimeType": mime}
                media = MediaFileUpload(image, mimetype=mime, resumable=True)
                created = (drive_service.files().create(body=image, media_body=media, fields="id,webViewLink").execute())
                file_permission = {"role": "reader", "type": "anyone"}
                drive_service.permissions().create(body=file_permission, fileId=created.get("id")).execute()
                link = created.get("webViewLink")
                full = 'https://drive.google.com/uc?export=view&id=' + link[link.find("d/")+2:link.rfind("/")]
                #self.ui.textBrowser_drive.append(full)
                list_link.append(full)
            self.ui.textBrowser_drive.append(" | ".join(list_link))
            self.ui.progressBar.setValue(100)
            QMessageBox.information(self, "?????????????? ",
                                    "???????????????? ???????? ??????????????????",
                                    QMessageBox.Ok)
            return file
        except (UnicodeDecodeError, OSError) as exe:
            print(f'?????????????????? ????????????: {exe}', traceback.format_exc())
            QMessageBox.critical(self, "???????????? ", "???????? ?? ???????????? ???? ???????????? ?????????????????? ?????????? ??????????????????!\n ?????? \n???????????????????????????????? ???????????? ??????????", QMessageBox.Ok)

    def create_exel(self, data):
        file, _ = QFileDialog.getOpenFileName(self, 'Open File', './',
                                               "Excel file (*.csv *.xlsx)")  # file ???????? ???? ??????????
        wb = openpyxl.load_workbook(file[0])
        ws = wb.worksheets[0]
        image_column = list(list(ws.values)[0]).index("ImageUrls")
        for row in range(2, 10):
            ws.cell(row=row, column=image_column+1).value = data
        wb.save(file)

    def create_analysis_exel(self, data):
        file, _ = QFileDialog.getOpenFileName(self, 'Open File', './',
                                               "Excel file (*.csv *.xlsx)")  # file ???????? ???? ??????????
        wb = openpyxl.load_workbook(file[0])
        ws = wb.worksheets[0]
        image_column = list(list(ws.values)[0]).index("ImageUrls")
        for row in range(2, 10):
            ws.cell(row=row, column=image_column+1).value = data
        wb.save(file)

    def dir_imageTo_exel_google(self):
        local_list_link = []
        directory = QFileDialog.getExistingDirectory(self, '?????????? ??????????', './')  # file ???????? ???? ??????????
        files = os.listdir(directory)
        self.ui.tableWidget_image_drive.setColumnCount(10)
        self.ui.tableWidget_image_drive.setRowCount(len(files))
        self.ui.tableWidget_image_drive.setIconSize(QtCore.QSize(100, 150))
        file, _ = QFileDialog.getOpenFileName(self, '?????????? Excel ??????????', './', "Excel file (*.csv *.xlsx)")  # file ???????? ???? ??????????
        wb = openpyxl.load_workbook(file)
        ws = wb.worksheets[0]
        image_column = list(list(ws.values)[0]).index("ImageUrls")
        self.ui.progressBar.setValue(20)
        all_image = 0
        for direct in files:
            adv = os.listdir(directory+"/"+direct)
            column = 0
            self.ui.progressBar.setValue(int(f"{column}0"))
            for image in adv:
                if image:
                    dir_image = directory + "/" + direct + "/" + image
                    item = QTableWidgetItem()
                    self.ui.tableWidget_image_drive.setItem(int(direct)-1, column, item)
                    item.setIcon(QtGui.QIcon(dir_image))
                    log.info(f"???????? ??????????????!")
                    drive_service = build('drive', 'v3', credentials=credentials)
                    basename = os.path.basename(dir_image)
                    mime = magic.from_file(dir_image, mime=True)
                    file_metadata = {"name": basename, "mimeType": mime}
                    media = MediaFileUpload(dir_image, mimetype=mime, resumable=True)
                    created = (drive_service.files().create(body=dir_image, media_body=media,
                                                                fields="id,webViewLink").execute())
                    file_permission = {"role": "reader", "type": "anyone"}
                    drive_service.permissions().create(body=file_permission, fileId=created.get("id")).execute()
                    link = created.get("webViewLink")
                    full = 'https://drive.google.com/uc?export=view&id=' + link[link.find("d/") + 2:link.rfind("/")]
                    local_list_link.append(full)
                    column += 1
            all_image += len(adv)
            self.ui.textBrowser_drive.append(f"{int(direct)}:   {' | '.join(local_list_link)}\n")
            ws.cell(row=int(direct) + 1, column=image_column + 1).value = str(" | ".join(local_list_link))
            local_list_link.clear()
        wb.save(file)
        self.ui.label_image_drive.setText(f"?????????????? ???????????????? : {all_image}")
        self.ui.progressBar.setValue(100)
        QMessageBox.information(self, "?????????????? ",
                                "???????????????? ???????? ??????????????????",
                                QMessageBox.Ok)
        self.ui.tableWidget_image_drive.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.ui.tableWidget_image_drive.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.ui.tableWidget_image_drive.horizontalHeader().setMinimumSectionSize(0)


    def analysis(self):
        driver = self.browser.start_browser(None)
        driver.get('https://avito.ru/')
        time.sleep(5)
        driver.refresh()
        time.sleep(3)
        self.ui.progressBar.setValue(10)
        if self.ui.radioButton_search.isChecked():
            # ????????????
            self.change_region(driver)
            # ???????????????? ????????????
            self.start_search(driver)
            self.ui.progressBar.setValue(20)
        else:
            driver.get(self.ui.lineEdit_anal_url.text())
        time.sleep(5)
        all_ad = driver.find_element(By.CLASS_NAME, "page-title-count-wQ7pG").text
        self.data['?????????? ????????????????????'] = all_ad
        #item = driver.find_elements(By.XPATH, "//div[@class='iva-item-root-_lk9K photo-slider-slider-S15A_ iva-item-list-rfgcH iva-item-redesign-rop6P iva-item-responsive-_lbhG items-item-My3ih items-listItem-Gd1jN js-catalog-item-enum']")
        action = ActionChains(driver)
        item = driver.find_elements(By.XPATH, "//div[@itemtype='http://schema.org/Product']")
        print(len(item))
        self.ui.progressBar.setValue(40)
        for i in item:
            time.sleep(0.5)
            try:
                up = i.find_element(By.CLASS_NAME, "styles-arrow-jfRdd")
                action.move_to_element(up).perform()
                elements = i.find_elements(By.XPATH, "//img[@class='style-image-wPviB']")
                print(21, len(elements))
                for element in elements:
                    upx = element.get_attribute("src")
                    value = upx[upx.rfind("/") + 1: upx.rfind(".")]
                    if value.rfind("_") != -1:
                        value = value[:value.rfind("_")]
                    self.data[value] = self.data[value] + 1
            except NoSuchElementException:
                #self.data['?????? ????????????????'] = (self.data['?????? ????????????????'])+1
                continue
        self.ui.progressBar.setValue(100)
        QMessageBox.information(self, "?????????????? ",
                                "???????????? '???????????? ?????????????? ??????????' ??????????????????",
                                QMessageBox.Ok)
        self.write_tablewidget(self.data)


    def analysis_competitor(self):
        json_com = []
        login = self.ui.comboBox_login.currentText()
        passw = self.cursor.execute(f"SELECT * FROM  db_profile WHERE number = '{login}'").fetchone()[3]
        driver = self.browser.authorization(login, passw)
        time.sleep(3)
        self.ui.progressBar.setValue(10)
        if self.ui.radioButton_search.isChecked():
            # ????????????
            self.change_region(driver)
            # ???????????????? ????????????
            self.start_search(driver)
            self.ui.progressBar.setValue(20)
        else:
            driver.get(self.ui.lineEdit_anal_url.text())
        time.sleep(5)
        action = ActionChains(driver)
        all_ad = driver.find_element(By.CLASS_NAME, "page-title-count-wQ7pG").text
        self.data['?????????? ????????????????????'] = all_ad
        pagination = driver.find_elements(By.XPATH, f"//span[@class='pagination-item-JJq_j']")
        if len(pagination) == 0:
            pagi = 3
        else:
            pagi = int(pagination[-1].text) - 1
        print(f"?????????? ?????????????? {pagi}")
        i = 2
        print(f"???????????????? {i-1}")
        while i < pagi:
            self.ui.progressBar.setValue(int(f"{i}0"))
            time.sleep(3)
            item = driver.find_elements(By.XPATH,
                                        "//div[@class='iva-item-root-_lk9K photo-slider-slider-S15A_ iva-item-list-rfgcH iva-item-redesign-rop6P iva-item-responsive-_lbhG items-item-My3ih items-listItem-Gd1jN js-catalog-item-enum']")
            #print("?????????? ??????????????", len(item))
            len_item = len(item)
            print("?????????????? ?????????????????? ???? ???????????????? :", len_item)
            if len_item > int(all_ad):
                item = item[:int(all_ad)]
            print("???? ???????????????? ???????????????? :", len(item))
            j = 1
            for el in item:
                dict_com = {}
                upper = []
                time.sleep(0.5)
                try:
                    up = el.find_element(By.CLASS_NAME, "styles-arrow-jfRdd")
                    action.move_to_element(up).perform()
                    elements = el.find_elements(By.XPATH, "//img[@class='style-image-wPviB']")
                    for element in elements:
                        upx = element.get_attribute("src")
                        value = upx[upx.rfind("/") + 1: upx.rfind(".")]
                        if value.rfind("_") != -1:
                            value = value[:value.rfind("_")]
                        upper.append(config.price_ad.get(value))
                except NoSuchElementException:
                    upper.append("?????? ?????????????? ??????????")
                ad = el.find_element(By.CLASS_NAME, "iva-item-sliderLink-uLz1v")
                action.move_to_element(ad).click().perform()
                driver.switch_to.window(driver.window_handles[1])
                WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((By.XPATH, "//span[@class='title-info-title-text']")))
                title = driver.find_element(By.XPATH, "//span[@class='title-info-title-text']").text
                dict_com['??????????????????'] = title
                try:
                    price = driver.find_element(By.XPATH, "//span[@class='js-item-price style-item-price-text-2u_qK text-text-1PdBw text-size-xxl-1Uoae']").text
                    dict_com['????????'] = price
                except NoSuchElementException:
                    dict_com['????????'] = '?????? ????????????'

                dict_com['?????????????? ????????????'] = ",".join(upper)

                try:
                    views = driver.find_element(By.XPATH, "//span[@data-marker='item-view/total-views']").text
                    dict_com['???????????????????? ??????????'] = views[:views.find(" ")]
                except NoSuchElementException:
                    dict_com['???????????????????? ??????????'] = '?????? ????????????'

                try:
                    views = driver.find_element(By.XPATH, "//span[@data-marker='item-view/today-views']").text
                    x = re.findall('[0-9]+',  views)
                    dict_com['???????????????????? ??????????????'] = x[0]
                except NoSuchElementException:
                    dict_com['???????????????????? ??????????'] = '?????? ????????????'
                try:
                    time_up = driver.find_element(By.XPATH, "//span[@data-marker='item-view/item-date']").text
                    dict_com['?????????? ????????????????????'] = time_up[1:]
                except NoSuchElementException:
                    dict_com['?????????? ????????????????????'] = '?????? ????????????'

                try:
                    adress = driver.find_element(By.XPATH, "//span[@class='style-item-address__string-3Ct0s']").text
                    dict_com['??????????'] = adress
                except NoSuchElementException:
                    dict_com['??????????'] = '?????? ????????????'

                try:
                    id = driver.find_element(By.XPATH, "//span[@data-marker='item-view/item-id']").text
                    dict_com['ID'] = id
                except NoSuchElementException:
                    dict_com['ID'] = '?????? ????????????'

                try:
                    description = driver.find_element(By.XPATH, "//div[@data-marker='item-view/item-description']").text
                    dict_com['??????????'] = description
                except NoSuchElementException:
                    dict_com['??????????'] = '?????? ????????????'

                try:
                    dict_com['???????????????? ?? ????????????'] = len(description)
                except NoSuchElementException:
                    dict_com['???????????????? ?? ????????????'] = '?????? ????????????'

                try:
                    name = driver.find_element(By.XPATH, "//span[@class='text-text-1PdBw text-size-ms-23YBR']").text
                    dict_com['?????? ????????????????????'] = name
                except NoSuchElementException:
                    dict_com['?????? ????????????????????'] = '?????? ????????????'

                try:
                    self.status = driver.find_element(By.XPATH, "//div[@data-marker='seller-info/label']").text
                    dict_com['???????????? ????????????????????'] = self.status
                except NoSuchElementException:
                    dict_com['???????????? ????????????????????'] = '?????? ??????????????'

                try:
                    stars = driver.find_element(By.XPATH, "//span[@class='style-seller-info-rating-score-KA-Kw']").text
                    dict_com['??????. ??????????'] = stars
                except NoSuchElementException:
                    dict_com['??????. ??????????'] = '?????? ??????????'

                try:
                    review = driver.find_element(By.XPATH, "//span[@data-marker='rating-caption/rating']").text
                    dict_com['??????. ??????????????'] = review
                except NoSuchElementException:
                    dict_com['??????. ??????????????'] = '?????? ??????????????'

                try:
                    all_item_user = 0
                    all_item_user = driver.find_element(By.XPATH, "//a[@class='button-button-2Fo5k button-size-s-3-rn6 button-default-mSfac width-width-12-2VZLz']").text
                except NoSuchElementException:
                    pass


                dict_com['???????????? ???? ????????????????????'] = driver.current_url

                com = driver.find_element(By.XPATH, "//a[@class='link-link-39EVK link-design-default-2sPEv link-novisited-1w4JY style-seller-name-link-1Ms2R']")
                action.move_to_element(com).click().perform()
                WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((By.XPATH, "//span[@class='desktop-1r4tu1s']")))
                try:
                    data_reg = driver.find_element(By.XPATH, "//span[@data-marker='registered']")
                    dict_com['???????? ??????????????????????'] = data_reg.text
                except NoSuchElementException:
                    dict_com['???????? ??????????????????????'] = "?????? ????????"

                try:
                    if all_item_user != 0:
                        z = re.findall('[0-9]+', all_item_user)
                        dict_com['????????????????????'] = z[0]
                    else:
                        if self.status == "???????????????????? ????????":
                            do_profile_adver = driver.find_elements(By.XPATH, "//div[@class='styles-root-_gnXE photo-slider-slider-S15A_  styles-responsive-m3Vnz']")
                            dict_com['????????????????????'] = len(do_profile_adver)
                        if self.status == "????????????????":
                            do_company_item = driver.find_elements(By.XPATH, "//span[@class='desktop-1r4tu1s']")
                            #print(len(do_company_item))
                            #for g in do_company_item:
                                #print(g.text)
                            if len(do_company_item) > 1:
                                z = re.findall('[0-9]+', do_company_item[-1].text)
                                dict_com['????????????????????'] = z[0]
                            else:
                                do_company_adver = driver.find_elements(By.XPATH, "//div[@class='Category-item-vzcCj']")
                                dict_com['????????????????????'] = len(do_company_adver)
                except NoSuchElementException:
                    dict_com['????????????????????'] = "?????? ????????????"
                dict_com['???????????? ???? ??????????????'] = driver.current_url
                time.sleep(0.1)
                print(f"???????????????????? {j}")
                print("_______________________")
                j+=1
                json_com.append(dict_com)
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                time.sleep(1)
            if len(pagination) != 0:
                st = driver.find_element(By.XPATH, f"//span[@data-marker='page({i})']")
                action.move_to_element(st).click().perform()
            else:
                break
            i+=1
            time.sleep(2)
        print(json_com)
        da = pd.DataFrame(json_com)
        da.to_excel('DATAFILE.xlsx')
        self.ui.progressBar.setValue(100)
        QMessageBox.information(self, "?????????????? ",
                             "???????????? '???????????? ??????????????????????' ??????????????????",
                             QMessageBox.Ok)



    def change_region(self, driver):
        driver.find_element(By.XPATH, "//div[@class='main-select-JJyaZ main-location-XUs1_']").click()
        time.sleep(0.2)
        region = driver.find_element(By.CLASS_NAME, "suggest-input-rORJM")
        region.send_keys(self.ui.lineEdit_anal_region.text()[:2])
        region.send_keys(self.ui.lineEdit_anal_region.text()[2:])
        time.sleep(0.4)
        #self.change.show()
        region.send_keys(Keys.DOWN)
        region.send_keys(Keys.ENTER)
        time.sleep(0.4)
        driver.find_element(By.XPATH, "//button[@class='button-button-CmK9a button-size-m-LzYrF button-primary-x_x8w']").click()

    # ????????????????
    def start_search(self, driver):
        list_category = []
        search = self.ui.lineEdit_anal_nisha.text()
        self.data['???????????????? ????????????'] = search
        time.sleep(0.2)
        search_from = driver.find_element(By.CLASS_NAME, "input-input-Zpzc1")
        search_from.send_keys(search[:1])
        search_from.send_keys(search[1:])
        time.sleep(0.5)
        item_category = driver.find_elements(By.XPATH, "//span[@class='suggest-itemDescription-oibjd text-text-LurtD text-size-s-BxGpL']")
        #print(len(item_category))
        for item in item_category:
            try:
                list_category.append(item.text)
                print(item.text)
            except NoSuchElementException:
                break
        self.change.set_combobox(list_category)
        self.change.show()
        item = self.change.return_combobox()
        print(item)
        #category = driver.find_element(By.XPATH, "//span[@class='suggest-itemDescription-_9rnO text-text-LurtD text-size-s-BxGpL']")
        #self.data['??????????????????'] = category.text
        #print(category.text)

        search_from.send_keys(Keys.DOWN)
        time.sleep(0.1)
        search_from.send_keys(Keys.ENTER)
        time.sleep(2)

    def write_tablewidget(self, data):
        self.ui.tableWidget_analyz.setColumnCount(11)
        self.ui.tableWidget_analyz.setHorizontalHeaderLabels(
            ['???????????????? ????????????', '??????????????????', '?????????? ????????????????????', '????????????????', '???????????? ???? XL', 'XL',  'x20', 'x15', 'x10','x5', 'x2'])
        records = list(data.values())
        records = [str(item) for item in records]
        self.ui.tableWidget_analyz.setRowCount(1)
        for column in range(11):
            self.ui.tableWidget_analyz.setItem(0, column, QTableWidgetItem(records[column]))
        self.ui.tableWidget_analyz.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.ui.tableWidget_analyz.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.ui.tableWidget_analyz.horizontalHeader().setMinimumSectionSize(0)